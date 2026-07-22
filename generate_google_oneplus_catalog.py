import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

google_oneplus_products = [
    # ------------------- GOOGLE PHONES -------------------
    {
        "name": "Google Pixel 9 Pro XL (2024)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 2900,
        "colors": "Obsidian, Porcelain, Hazel, Rose Quartz",
        "variants": [
            {"spec": "128GB / 16GB RAM", "boost": 0},
            {"spec": "256GB / 16GB RAM", "boost": 300},
            {"spec": "512GB / 16GB RAM", "boost": 700},
            {"spec": "1TB / 16GB RAM", "boost": 1200}
        ],
        "shortDescription": "Sell your used Google Pixel 9 Pro XL for instant cash in Dubai & UAE. Features Google Tensor G4 processor, 16GB RAM, built-in Gemini Advanced AI, 6.8-inch Super Actua OLED display, and pro-grade 50MP triple camera with 30x Super Res Zoom."
    },
    {
        "name": "Google Pixel 9 Pro (2024)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 2500,
        "colors": "Obsidian, Porcelain, Hazel, Rose Quartz",
        "variants": [
            {"spec": "128GB / 16GB RAM", "boost": 0},
            {"spec": "256GB / 16GB RAM", "boost": 250},
            {"spec": "512GB / 16GB RAM", "boost": 600}
        ],
        "shortDescription": "Get the best trade-in value for Google Pixel 9 Pro in UAE. Powered by Google Tensor G4 with 16GB RAM, compact 6.3-inch Super Actua LTPO display, 50MP triple camera system, and 7 years of Pixel OS & security updates."
    },
    {
        "name": "Google Pixel 9 (2024)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 1900,
        "colors": "Obsidian, Porcelain, Wintergreen, Peony",
        "variants": [
            {"spec": "128GB / 12GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 250}
        ],
        "shortDescription": "Instant cash for Google Pixel 9 in Dubai. Packed with Tensor G4 chip, 12GB RAM, 6.3-inch 120Hz Actua OLED screen, upgraded 50MP main + 48MP ultrawide cameras, and advanced Gemini AI photography features."
    },
    {
        "name": "Google Pixel 9 Pro Fold (2024)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 3800,
        "colors": "Obsidian, Porcelain",
        "variants": [
            {"spec": "256GB / 16GB RAM", "boost": 0},
            {"spec": "512GB / 16GB RAM", "boost": 500}
        ],
        "shortDescription": "Sell your Google Pixel 9 Pro Fold online in UAE. Ultra-thin foldable smartphone featuring a huge 8.0-inch Super Actua Flex inner display, 6.3-inch cover screen, Tensor G4, 16GB RAM, and IPX8 water resistance."
    },
    {
        "name": "Google Pixel 8a (2024)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 1200,
        "colors": "Obsidian, Porcelain, Bay, Aloe",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 180}
        ],
        "shortDescription": "Trade in Google Pixel 8a for cash in Abu Dhabi & Dubai. Powered by Google Tensor G3, 6.1-inch 120Hz Actua OLED screen, 64MP dual rear cameras with Best Take & Magic Eraser, and 7 years of security support."
    },
    {
        "name": "Google Pixel 8 Pro (2023)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 1800,
        "colors": "Obsidian, Porcelain, Bay, Mint",
        "variants": [
            {"spec": "128GB / 12GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 200},
            {"spec": "512GB / 12GB RAM", "boost": 450}
        ],
        "shortDescription": "Sell Google Pixel 8 Pro with free doorstep pickup in Dubai. Features Tensor G3 chip, 6.7-inch Super Actua LTPO display, built-in Temperature Sensor, Audio Magic Eraser, and pro-level 50MP camera with 5x optical zoom."
    },
    {
        "name": "Google Pixel 8 (2023)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 1350,
        "colors": "Obsidian, Hazel, Rose, Mint",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 180}
        ],
        "shortDescription": "Get instant buyback quote for Google Pixel 8 in UAE. Compact flagship with Google Tensor G3, 6.2-inch 120Hz Actua display, 50MP camera with Macro Focus, and 4575mAh fast-charging battery."
    },
    {
        "name": "Google Pixel Fold (2023)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 2400,
        "colors": "Obsidian, Porcelain",
        "variants": [
            {"spec": "256GB / 12GB RAM", "boost": 0},
            {"spec": "512GB / 12GB RAM", "boost": 400}
        ],
        "shortDescription": "Sell your Google Pixel Fold in Dubai. Google's first foldable phone with a 7.6-inch OLED inner display, Tensor G2 processor, 12GB RAM, multi-tasking Split Screen, and triple rear camera setup."
    },
    {
        "name": "Google Pixel 7a (2023)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 850,
        "colors": "Charcoal, Snow, Sea, Coral",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0}
        ],
        "shortDescription": "Resale price valuation for Google Pixel 7a in UAE. Equipped with Tensor G2, 6.1-inch 90Hz OLED display, wireless charging support, 64MP dual camera system, and Titan M2 security coprocessor."
    },
    {
        "name": "Google Pixel 7 Pro (2022)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 1200,
        "colors": "Obsidian, Snow, Hazel",
        "variants": [
            {"spec": "128GB / 12GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 150},
            {"spec": "512GB / 12GB RAM", "boost": 350}
        ],
        "shortDescription": "Sell used Google Pixel 7 Pro for top cash in Dubai. Premium phone with 6.7-inch QHD+ 120Hz LTPO display, Tensor G2 chip, 50MP triple camera with 30x Super Res Zoom, and Macro Focus."
    },
    {
        "name": "Google Pixel 7 (2022)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 950,
        "colors": "Obsidian, Snow, Lemongrass",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 150}
        ],
        "shortDescription": "Instant sell offer for Google Pixel 7 in UAE. Sleek matte aluminum enclosure, 6.3-inch 90Hz OLED screen, Tensor G2 chip, Real Tone photography, and Cinematic Blur video mode."
    },
    {
        "name": "Google Pixel 6 Pro (2021)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 800,
        "colors": "Stormy Black, Cloudy White, Sorta Sunny",
        "variants": [
            {"spec": "128GB / 12GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 120}
        ],
        "shortDescription": "Sell Google Pixel 6 Pro in UAE. Powered by the first-gen Google Tensor chip, 6.7-inch QHD+ 120Hz LTPO display, 50MP main + 48MP telephoto camera, and 5003mAh long-lasting battery."
    },
    {
        "name": "Google Pixel 6 (2021)",
        "brand": "Google",
        "category": "mobile",
        "basePrice": 650,
        "colors": "Stormy Black, Sorta Seafoam, Kinda Coral",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 100}
        ],
        "shortDescription": "Trade in Google Pixel 6 for instant cash in Dubai. Features custom Google Tensor SoC, dual-tone Gorilla Glass Victus design, 6.4-inch 90Hz OLED screen, and Magic Eraser."
    },

    # ------------------- ONEPLUS PHONES -------------------
    {
        "name": "OnePlus 12 (2024)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 2300,
        "colors": "Silky Black, Flowy Emerald",
        "variants": [
            {"spec": "256GB / 12GB RAM", "boost": 0},
            {"spec": "512GB / 16GB RAM", "boost": 350},
            {"spec": "1TB / 16GB RAM", "boost": 700}
        ],
        "shortDescription": "Sell your used OnePlus 12 for instant cash in Dubai & UAE. Powered by Qualcomm Snapdragon 8 Gen 3, 4th Gen Hasselblad Camera for Mobile, 6.82-inch 2K 120Hz ProXDR display, 5400mAh battery, and 100W SUPERVOOC charging."
    },
    {
        "name": "OnePlus 12R (2024)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 1450,
        "colors": "Cool Blue, Iron Gray",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 16GB RAM", "boost": 250}
        ],
        "shortDescription": "Get best trade-in value for OnePlus 12R in UAE. Performance flagship featuring Snapdragon 8 Gen 2, 6.78-inch 1.5K 120Hz LTPO 4.0 AMOLED screen, massive 5500mAh battery, and 100W fast charging."
    },
    {
        "name": "OnePlus Open (2023)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 3600,
        "colors": "Voyager Black, Emerald Dusk",
        "variants": [
            {"spec": "512GB / 16GB RAM", "boost": 0}
        ],
        "shortDescription": "Sell your OnePlus Open foldable smartphone for top cash in Dubai. Features Snapdragon 8 Gen 2, 7.82-inch 2K 120Hz Flexi-fluid inner display, Hasselblad Triple Main Camera system, and lightweight titanium alloy hinge."
    },
    {
        "name": "OnePlus 11 5G (2023)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 1500,
        "colors": "Titan Black, Eternal Green",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 16GB RAM", "boost": 250}
        ],
        "shortDescription": "Instant sell quote for OnePlus 11 5G in UAE. Powered by Snapdragon 8 Gen 2, 3rd Gen Hasselblad Camera, 6.7-inch 2K 120Hz Super Fluid AMOLED, Cryo-velocity cooling, and 100W SUPERVOOC."
    },
    {
        "name": "OnePlus 11R 5G (2023)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 1100,
        "colors": "Sonic Black, Galactic Silver",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 16GB RAM", "boost": 200}
        ],
        "shortDescription": "Sell OnePlus 11R 5G for cash in Abu Dhabi & Dubai. Packed with Snapdragon 8+ Gen 1 processor, 6.74-inch 120Hz Super Fluid AMOLED, 50MP Sony IMX890 main camera with OIS, and 100W fast charging."
    },
    {
        "name": "OnePlus 10 Pro 5G (2022)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 1150,
        "colors": "Volcanic Black, Emerald Forest",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 200}
        ],
        "shortDescription": "Trade in OnePlus 10 Pro 5G for instant cash in Dubai. Premium smartphone with Snapdragon 8 Gen 1, 2nd Gen Hasselblad Camera, 6.7-inch QHD+ 120Hz Fluid AMOLED LTPO, and 80W SUPERVOOC."
    },
    {
        "name": "OnePlus 10T 5G (2022)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 1000,
        "colors": "Moonstone Black, Jade Green",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 16GB RAM", "boost": 200}
        ],
        "shortDescription": "Sell OnePlus 10T 5G in UAE with free pickup. Extreme speed flagship with Snapdragon 8+ Gen 1, 150W SUPERVOOC charging (1-100% in 19 mins), 6.7-inch 120Hz Fluid AMOLED display, and 3D Cooling System."
    },
    {
        "name": "OnePlus 9 Pro 5G (2021)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 850,
        "colors": "Morning Mist, Stellar Black, Pine Green",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 150}
        ],
        "shortDescription": "Resale cash value for OnePlus 9 Pro 5G in Dubai. Features Hasselblad Camera for Mobile, Snapdragon 888 5G, 6.7-inch QHD+ 120Hz Smart LTPO Fluid Display 2.0, and 65T Warp Charge."
    },
    {
        "name": "OnePlus 9 5G (2021)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 700,
        "colors": "Astral Black, Winter Mist, Arctic Sky",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 120}
        ],
        "shortDescription": "Sell OnePlus 9 5G online in UAE. Co-developed with Hasselblad, powered by Snapdragon 888, 6.55-inch 120Hz Fluid AMOLED screen, 50MP Ultra-Wide lens, and 65W Warp Charge."
    },
    {
        "name": "OnePlus Nord 4 (2024)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 1150,
        "colors": "Obsidian Midnight, Oasis Green, Mercurial Silver",
        "variants": [
            {"spec": "256GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 120},
            {"spec": "512GB / 16GB RAM", "boost": 280}
        ],
        "shortDescription": "Sell OnePlus Nord 4 in Dubai. Sleek metal unibody design, Snapdragon 7+ Gen 3, 6.74-inch 1.5K 120Hz AMOLED, 5500mAh battery with 100W SUPERVOOC, and AI Productivity Tools."
    },
    {
        "name": "OnePlus Nord CE 4 (2024)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 800,
        "colors": "Dark Chrome, Celadon Marble",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 120}
        ],
        "shortDescription": "Instant buyback valuation for OnePlus Nord CE 4 in UAE. Equipped with Snapdragon 7 Gen 3, 6.7-inch 120Hz FHD+ AMOLED display, 50MP Sony LYT-600 main camera with OIS, and 100W SUPERVOOC."
    },
    {
        "name": "OnePlus Nord 3 5G (2023)",
        "brand": "OnePlus",
        "category": "mobile",
        "basePrice": 750,
        "colors": "Misty Green, Tempest Gray",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 16GB RAM", "boost": 150}
        ],
        "shortDescription": "Trade in OnePlus Nord 3 5G for instant cash in UAE. Features MediaTek Dimensity 9000 flagship chip, 6.74-inch 120Hz Super Fluid AMOLED, 50MP Sony IMX890 camera, and 80W SUPERVOOC."
    }
]

# Generate Excel Workbook
wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=16, bold=True, color="10B981")
bold_font = Font(name="Calibri", size=10, bold=True)
price_font = Font(name="Calibri", size=10, bold=True, color="059669")
border_thin = Side(border_style="thin", color="CBD5E1")
cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

# Sheet 1: Master Catalog
ws1 = wb.active
ws1.title = "Google & OnePlus Master Catalog"
ws1.append(["Google Pixel & OnePlus Mobile Phones - Admin Panel Master Catalog"])
ws1.merge_cells("A1:G1")
ws1["A1"].font = title_font
ws1.append([])

headers1 = [
    "Product Name", "Brand", "Category", "Base Price (AED)", 
    "Color Options", "Specification Variants & Price Boosts", "SEO Short Description"
]
ws1.append(headers1)
for col_num, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx = 4
for p in google_oneplus_products:
    spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
    ws1.append([
        p["name"], p["brand"], p["category"], p["basePrice"],
        p["colors"], spec_list_str, p["shortDescription"]
    ])
    ws1.cell(row=row_idx, column=1).font = bold_font
    ws1.cell(row=row_idx, column=4).font = price_font
    ws1.cell(row=row_idx, column=4).number_format = '#,##0'
    for c in range(1, 8):
        cell = ws1.cell(row=row_idx, column=c)
        cell.border = cell_border
        if c in [2, 3]: cell.alignment = Alignment(horizontal="center")
        elif c == 4: cell.alignment = Alignment(horizontal="right")
        else: cell.alignment = Alignment(horizontal="left", wrap_text=True)
    row_idx += 1

# Sheet 2: Itemized Variants
ws2 = wb.create_sheet(title="Itemized Spec Variants")
ws2.append(["Google & OnePlus Storage & Buyback Valuations Breakdown"])
ws2.merge_cells("A1:H1")
ws2["A1"].font = title_font
ws2.append([])

headers2 = [
    "Product Model Name", "Brand", "Category", "Base Price (AED)",
    "Variant Spec / Storage & RAM", "Spec Boost (AED)", "Total Valuation (AED)", "Colors"
]
ws2.append(headers2)
for col_num, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx2 = 4
for p in google_oneplus_products:
    for v in p["variants"]:
        total_val = p["basePrice"] + v["boost"]
        ws2.append([
            p["name"], p["brand"], p["category"], p["basePrice"],
            v["spec"], v["boost"], total_val, p["colors"]
        ])
        ws2.cell(row=row_idx2, column=1).font = bold_font
        ws2.cell(row=row_idx2, column=5).font = bold_font
        ws2.cell(row=row_idx2, column=7).font = price_font
        for c in range(1, 9):
            cell = ws2.cell(row=row_idx2, column=c)
            cell.border = cell_border
            if c in [2, 3]: cell.alignment = Alignment(horizontal="center")
            elif c in [4, 6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            else: cell.alignment = Alignment(horizontal="left")
        row_idx2 += 1

for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

# Save files to workspace and artifact directories
dirs = [r"c:\Users\shant\sellphone.ae", r"C:\Users\shant\.gemini\antigravity-ide\brain\ceae0dbf-d5b8-4d55-b6c1-296d4da1afdd"]

for d in dirs:
    os.makedirs(d, exist_ok=True)
    wb.save(os.path.join(d, "Google_OnePlus_Phones_Catalog.xlsx"))
    
    # Write CSV 1: Main Catalog
    with open(os.path.join(d, "Google_OnePlus_Phones_Catalog.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers1)
        for p in google_oneplus_products:
            spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
            writer.writerow([p["name"], p["brand"], p["category"], p["basePrice"], p["colors"], spec_list_str, p["shortDescription"]])

    # Write CSV 2: Itemized Variants
    with open(os.path.join(d, "Google_OnePlus_Variants_Breakdown.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers2)
        for p in google_oneplus_products:
            for v in p["variants"]:
                writer.writerow([p["name"], p["brand"], p["category"], p["basePrice"], v["spec"], v["boost"], p["basePrice"] + v["boost"], p["colors"]])

print("Successfully generated Google & OnePlus Excel and CSV files!")
