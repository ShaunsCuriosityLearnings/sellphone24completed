import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

apple_watches = [
    {
        "name": "Apple Watch Ultra 2 (2023)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 2200,
        "colors": "Titanium (Natural)",
        "variants": [
            {"spec": "49mm Titanium Case (GPS + Cellular)", "boost": 0}
        ],
        "shortDescription": "49mm Titanium case, 3000 nits display, S9 SiP, Double Tap gesture, Precision Dual-Frequency GPS, 36hr battery."
    },
    {
        "name": "Apple Watch Ultra (2022)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 1700,
        "colors": "Titanium (Natural)",
        "variants": [
            {"spec": "49mm Titanium Case (GPS + Cellular)", "boost": 0}
        ],
        "shortDescription": "49mm Titanium case, 2000 nits display, Action Button, 100m water resistance, 36hr battery."
    },
    {
        "name": "Apple Watch Series 9 (2023)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 1100,
        "colors": "Midnight, Starlight, Silver, Pink, (PRODUCT)RED, Graphite, Gold",
        "variants": [
            {"spec": "41mm Aluminum (GPS)", "boost": 0},
            {"spec": "41mm Aluminum (GPS + Cellular)", "boost": 100},
            {"spec": "45mm Aluminum (GPS)", "boost": 120},
            {"spec": "45mm Aluminum (GPS + Cellular)", "boost": 220},
            {"spec": "41mm Stainless Steel (Cellular)", "boost": 350},
            {"spec": "45mm Stainless Steel (Cellular)", "boost": 450}
        ],
        "shortDescription": "S9 SiP, Double Tap gesture, 2000 nits Always-On Retina display, Blood Oxygen & ECG apps."
    },
    {
        "name": "Apple Watch Series 8 (2022)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 850,
        "colors": "Midnight, Starlight, Silver, (PRODUCT)RED, Graphite, Gold",
        "variants": [
            {"spec": "41mm Aluminum (GPS)", "boost": 0},
            {"spec": "41mm Aluminum (GPS + Cellular)", "boost": 80},
            {"spec": "45mm Aluminum (GPS)", "boost": 100},
            {"spec": "45mm Aluminum (GPS + Cellular)", "boost": 180},
            {"spec": "41mm Stainless Steel (Cellular)", "boost": 280},
            {"spec": "45mm Stainless Steel (Cellular)", "boost": 380}
        ],
        "shortDescription": "Temperature sensing for cycle tracking, Crash Detection, Always-On Retina display, IP6X dust resistant."
    },
    {
        "name": "Apple Watch Series 7 (2021)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 650,
        "colors": "Midnight, Starlight, Green, Blue, (PRODUCT)RED, Graphite, Gold",
        "variants": [
            {"spec": "41mm Aluminum (GPS)", "boost": 0},
            {"spec": "41mm Aluminum (GPS + Cellular)", "boost": 60},
            {"spec": "45mm Aluminum (GPS)", "boost": 80},
            {"spec": "45mm Aluminum (GPS + Cellular)", "boost": 150},
            {"spec": "45mm Stainless Steel (Cellular)", "boost": 280}
        ],
        "shortDescription": "Nearly 20% larger screen area than Series 6, faster charging, crack-resistant front crystal."
    },
    {
        "name": "Apple Watch Series 6 (2020)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 500,
        "colors": "Space Gray, Silver, Gold, Blue, (PRODUCT)RED, Graphite",
        "variants": [
            {"spec": "40mm Aluminum (GPS)", "boost": 0},
            {"spec": "40mm Aluminum (GPS + Cellular)", "boost": 50},
            {"spec": "44mm Aluminum (GPS)", "boost": 70},
            {"spec": "44mm Aluminum (GPS + Cellular)", "boost": 120}
        ],
        "shortDescription": "Blood Oxygen sensor & app, S6 System in Package, Always-On Altimeter."
    },
    {
        "name": "Apple Watch Series 5 (2019)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 380,
        "colors": "Space Gray, Silver, Gold",
        "variants": [
            {"spec": "40mm Aluminum (GPS)", "boost": 0},
            {"spec": "40mm Aluminum (GPS + Cellular)", "boost": 40},
            {"spec": "44mm Aluminum (GPS)", "boost": 60},
            {"spec": "44mm Aluminum (GPS + Cellular)", "boost": 100}
        ],
        "shortDescription": "First Always-On Retina display, Built-in Compass, International Emergency Calling."
    },
    {
        "name": "Apple Watch SE 2nd Gen (2022)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 550,
        "colors": "Midnight, Starlight, Silver",
        "variants": [
            {"spec": "40mm Aluminum (GPS)", "boost": 0},
            {"spec": "40mm Aluminum (GPS + Cellular)", "boost": 60},
            {"spec": "44mm Aluminum (GPS)", "boost": 80},
            {"spec": "44mm Aluminum (GPS + Cellular)", "boost": 140}
        ],
        "shortDescription": "S8 SiP (same chip as Series 8), Crash Detection, redesign nylon composite back case."
    },
    {
        "name": "Apple Watch SE 1st Gen (2020)",
        "brand": "Apple",
        "category": "smartwatches",
        "basePrice": 350,
        "colors": "Space Gray, Silver, Gold",
        "variants": [
            {"spec": "40mm Aluminum (GPS)", "boost": 0},
            {"spec": "40mm Aluminum (GPS + Cellular)", "boost": 40},
            {"spec": "44mm Aluminum (GPS)", "boost": 60},
            {"spec": "44mm Aluminum (GPS + Cellular)", "boost": 90}
        ],
        "shortDescription": "Essential features at a great value, S5 dual-core processor, Retina display, Fall Detection."
    }
]

# Generate Excel Workbook
wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=16, bold=True, color="10B981")
bold_font = Font(name="Calibri", size=10, bold=True)
price_font = Font(name="Calibri", size=10, bold=True, color="059669")
border_thin = Side(border_style="thin", color="CBD5E1")
cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

# Sheet 1: Master Catalog
ws1 = wb.active
ws1.title = "Apple Watch Admin Master Catalog"
ws1.append(["Apple Watch Smartwatches - Admin Panel Master Catalog"])
ws1.merge_cells("A1:G1")
ws1["A1"].font = title_font
ws1.append([])

headers1 = [
    "Product Name", "Brand", "Category", "Base Price (AED)", 
    "Color Options", "Specification Variants & Price Boosts", "Short Description"
]
ws1.append(headers1)
for col_num, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx = 4
for p in apple_watches:
    spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
    ws1.append([
        p["name"], p["brand"], p["category"], p["basePrice"],
        p["colors"], spec_list_str, p["shortDescription"]
    ])
    ws1.cell(row=row_idx, column=1).font = bold_font
    ws1.cell(row=row_idx, column=4).font = price_font
    ws1.cell(row=row_idx, column=4).number_format = '#,##0'
    for c in range(1, 8):
        cell = ws1.cell(row=row_idx, column=c)
        cell.border = cell_border
        if c in [2, 3]: cell.alignment = Alignment(horizontal="center")
        elif c == 4: cell.alignment = Alignment(horizontal="right")
        else: cell.alignment = Alignment(horizontal="left", wrap_text=True)
    row_idx += 1

# Sheet 2: Itemized Variants
ws2 = wb.create_sheet(title="Itemized Spec Variants")
ws2.append(["Apple Watch Specifications & Buyback Valuations Breakdown"])
ws2.merge_cells("A1:H1")
ws2["A1"].font = title_font
ws2.append([])

headers2 = [
    "Product Model Name", "Brand", "Category", "Base Price (AED)",
    "Variant Spec / Size & Connectivity", "Spec Boost (AED)", "Total Valuation (AED)", "Colors"
]
ws2.append(headers2)
for col_num, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx2 = 4
for p in apple_watches:
    for v in p["variants"]:
        total_val = p["basePrice"] + v["boost"]
        ws2.append([
            p["name"], p["brand"], p["category"], p["basePrice"],
            v["spec"], v["boost"], total_val, p["colors"]
        ])
        ws2.cell(row=row_idx2, column=1).font = bold_font
        ws2.cell(row=row_idx2, column=5).font = bold_font
        ws2.cell(row=row_idx2, column=7).font = price_font
        for c in range(1, 9):
            cell = ws2.cell(row=row_idx2, column=c)
            cell.border = cell_border
            if c in [2, 3]: cell.alignment = Alignment(horizontal="center")
            elif c in [4, 6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            else: cell.alignment = Alignment(horizontal="left")
        row_idx2 += 1

for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

# Save files to workspace and artifact directories
dirs = [r"c:\Users\shant\sellphone.ae", r"C:\Users\shant\.gemini\antigravity-ide\brain\ceae0dbf-d5b8-4d55-b6c1-296d4da1afdd"]

for d in dirs:
    os.makedirs(d, exist_ok=True)
    wb.save(os.path.join(d, "Apple_Watch_Catalog.xlsx"))
    
    # Write CSV 1: Main Catalog
    with open(os.path.join(d, "Apple_Watch_Catalog.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers1)
        for p in apple_watches:
            spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
            writer.writerow([p["name"], p["brand"], p["category"], p["basePrice"], p["colors"], spec_list_str, p["shortDescription"]])

    # Write CSV 2: Itemized Variants
    with open(os.path.join(d, "Apple_Watch_Variants_Breakdown.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers2)
        for p in apple_watches:
            for v in p["variants"]:
                writer.writerow([p["name"], p["brand"], p["category"], p["basePrice"], v["spec"], v["boost"], p["basePrice"] + v["boost"], p["colors"]])

print("Successfully generated Apple Watch Excel and CSV files!")
