import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

samsung_products = [
    # 1. Galaxy S-Series (Flagships)
    {
        "name": "Samsung Galaxy S24 Ultra (2024)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 3400,
        "colors": "Titanium Gray, Titanium Black, Titanium Violet, Titanium Yellow",
        "variants": [
            {"spec": "256GB / 12GB RAM", "boost": 0},
            {"spec": "512GB / 12GB RAM", "boost": 350},
            {"spec": "1TB / 12GB RAM", "boost": 750},
        ],
        "shortDescription": "Snapdragon 8 Gen 3 for Galaxy, 200MP camera, 6.8\" QHD+ Dynamic AMOLED 2X, S Pen."
    },
    {
        "name": "Samsung Galaxy S24+ (2024)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 2500,
        "colors": "Onyx Black, Marble Gray, Cobalt Violet, Amber Yellow",
        "variants": [
            {"spec": "256GB / 12GB RAM", "boost": 0},
            {"spec": "512GB / 12GB RAM", "boost": 300},
        ],
        "shortDescription": "6.7\" QHD+ Dynamic AMOLED 2X display, 4900mAh battery, Galaxy AI features."
    },
    {
        "name": "Samsung Galaxy S24 (2024)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 1900,
        "colors": "Onyx Black, Marble Gray, Cobalt Violet, Amber Yellow",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 200},
            {"spec": "512GB / 8GB RAM", "boost": 400},
        ],
        "shortDescription": "6.2\" FHD+ Dynamic AMOLED 2X display, 4000mAh battery, Galaxy AI features."
    },
    {
        "name": "Samsung Galaxy S23 Ultra (2023)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 2500,
        "colors": "Phantom Black, Green, Lavender, Cream",
        "variants": [
            {"spec": "256GB / 8GB RAM", "boost": 0},
            {"spec": "512GB / 12GB RAM", "boost": 300},
            {"spec": "1TB / 12GB RAM", "boost": 650},
        ],
        "shortDescription": "Snapdragon 8 Gen 2 for Galaxy, 200MP camera, 100x Space Zoom, S Pen."
    },
    {
        "name": "Samsung Galaxy S23+ (2023)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 1800,
        "colors": "Phantom Black, Cream, Green, Lavender",
        "variants": [
            {"spec": "256GB / 8GB RAM", "boost": 0},
            {"spec": "512GB / 8GB RAM", "boost": 250},
        ],
        "shortDescription": "6.6\" Dynamic AMOLED 2X, 4700mAh battery, 50MP triple camera system."
    },
    {
        "name": "Samsung Galaxy S23 (2023)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 1400,
        "colors": "Phantom Black, Cream, Green, Lavender",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 150},
            {"spec": "512GB / 8GB RAM", "boost": 300},
        ],
        "shortDescription": "Compact 6.1\" Dynamic AMOLED 2X display, Snapdragon 8 Gen 2 processor."
    },
    {
        "name": "Samsung Galaxy S22 Ultra (2022)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 1750,
        "colors": "Phantom Black, Phantom White, Burgundy, Green",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 12GB RAM", "boost": 200},
            {"spec": "512GB / 12GB RAM", "boost": 400},
            {"spec": "1TB / 12GB RAM", "boost": 700},
        ],
        "shortDescription": "Built-in S Pen, 108MP main camera, 6.8\" 120Hz AMOLED display."
    },
    {
        "name": "Samsung Galaxy S22+ (2022)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 1250,
        "colors": "Phantom Black, Phantom White, Green, Pink Gold",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 150},
        ],
        "shortDescription": "6.6\" Dynamic AMOLED 2X, Armor Aluminum frame, 4500mAh battery."
    },
    {
        "name": "Samsung Galaxy S22 (2022)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 950,
        "colors": "Phantom Black, Phantom White, Green, Pink Gold",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 120},
        ],
        "shortDescription": "6.1\" Dynamic AMOLED 2X, Nightography 50MP camera."
    },
    {
        "name": "Samsung Galaxy S21 FE 5G (2022)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 750,
        "colors": "Olive, Lavender, White, Graphite",
        "variants": [
            {"spec": "128GB / 6GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 100},
        ],
        "shortDescription": "Fan Edition flagship with 6.4\" 120Hz AMOLED, 32MP selfie camera."
    },

    # 2. Galaxy Foldables (Z Series)
    {
        "name": "Samsung Galaxy Z Fold 6 (2024)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 4200,
        "colors": "Silver Shadow, Pink, Navy",
        "variants": [
            {"spec": "256GB / 12GB RAM", "boost": 0},
            {"spec": "512GB / 12GB RAM", "boost": 450},
            {"spec": "1TB / 12GB RAM", "boost": 900},
        ],
        "shortDescription": "Dual AMOLED 120Hz displays, 7.6\" Main display, Snapdragon 8 Gen 3, S Pen support."
    },
    {
        "name": "Samsung Galaxy Z Flip 6 (2024)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 2700,
        "colors": "Silver Shadow, Yellow, Blue, Mint",
        "variants": [
            {"spec": "256GB / 12GB RAM", "boost": 0},
            {"spec": "512GB / 12GB RAM", "boost": 350},
        ],
        "shortDescription": "3.4\" Flex Window cover screen, 50MP camera, FlexCam AI zoom."
    },
    {
        "name": "Samsung Galaxy Z Fold 5 (2023)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 3100,
        "colors": "Icy Blue, Phantom Black, Cream",
        "variants": [
            {"spec": "256GB / 12GB RAM", "boost": 0},
            {"spec": "512GB / 12GB RAM", "boost": 350},
            {"spec": "1TB / 12GB RAM", "boost": 750},
        ],
        "shortDescription": "Zero-gap Flex Hinge, 7.6\" Dynamic AMOLED 2X, Snapdragon 8 Gen 2."
    },
    {
        "name": "Samsung Galaxy Z Flip 5 (2023)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 1800,
        "colors": "Mint, Graphite, Cream, Lavender",
        "variants": [
            {"spec": "256GB / 8GB RAM", "boost": 0},
            {"spec": "512GB / 8GB RAM", "boost": 250},
        ],
        "shortDescription": "Large Flex Window, hands-free Flex Mode selfies, IPX8 water resistance."
    },
    {
        "name": "Samsung Galaxy Z Fold 4 (2022)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 2200,
        "colors": "Graygreen, Phantom Black, Beige",
        "variants": [
            {"spec": "256GB / 12GB RAM", "boost": 0},
            {"spec": "512GB / 12GB RAM", "boost": 250},
            {"spec": "1TB / 12GB RAM", "boost": 550},
        ],
        "shortDescription": "50MP camera system, Under Display Camera, Taskbar multi-tasking."
    },

    # 3. Galaxy A-Series (Popular Mid-Range)
    {
        "name": "Samsung Galaxy A55 5G (2024)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 950,
        "colors": "Awesome Iceblue, Awesome Lilac, Awesome Lemon, Awesome Navy",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 120},
            {"spec": "256GB / 12GB RAM", "boost": 220},
        ],
        "shortDescription": "Metal frame, 50MP OIS camera, 6.6\" Super AMOLED 120Hz display."
    },
    {
        "name": "Samsung Galaxy A35 5G (2024)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 720,
        "colors": "Awesome Iceblue, Awesome Lilac, Awesome Lemon, Awesome Navy",
        "variants": [
            {"spec": "128GB / 6GB RAM", "boost": 0},
            {"spec": "128GB / 8GB RAM", "boost": 80},
            {"spec": "256GB / 8GB RAM", "boost": 150},
        ],
        "shortDescription": "Glass back, 50MP main camera, 5000mAh battery, IP67 rating."
    },
    {
        "name": "Samsung Galaxy A54 5G (2023)",
        "brand": "Samsung",
        "category": "mobile",
        "basePrice": 750,
        "colors": "Awesome Lime, Awesome Graphite, Awesome Violet, Awesome White",
        "variants": [
            {"spec": "128GB / 8GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 100},
        ],
        "shortDescription": "Premium glass back design, Nightography camera, 6.4\" 120Hz Super AMOLED."
    },

    # 4. Galaxy Tablets
    {
        "name": "Samsung Galaxy Tab S9 Ultra (2023)",
        "brand": "Samsung",
        "category": "tablets",
        "basePrice": 2900,
        "colors": "Beige, Graphite",
        "variants": [
            {"spec": "256GB / 12GB RAM (Wi-Fi)", "boost": 0},
            {"spec": "512GB / 12GB RAM (5G)", "boost": 400},
            {"spec": "1TB / 16GB RAM (5G)", "boost": 900},
        ],
        "shortDescription": "Massive 14.6\" Dynamic AMOLED 2X, S Pen included, IP68 water resistance."
    },
    {
        "name": "Samsung Galaxy Tab S9+ (2023)",
        "brand": "Samsung",
        "category": "tablets",
        "basePrice": 2300,
        "colors": "Beige, Graphite",
        "variants": [
            {"spec": "256GB / 12GB RAM (Wi-Fi)", "boost": 0},
            {"spec": "512GB / 12GB RAM (5G)", "boost": 350},
        ],
        "shortDescription": "12.4\" Dynamic AMOLED 2X, Snapdragon 8 Gen 2 for Galaxy."
    },
    {
        "name": "Samsung Galaxy Tab S9 (2023)",
        "brand": "Samsung",
        "category": "tablets",
        "basePrice": 1800,
        "colors": "Beige, Graphite",
        "variants": [
            {"spec": "128GB / 8GB RAM (Wi-Fi)", "boost": 0},
            {"spec": "256GB / 12GB RAM (5G)", "boost": 300},
        ],
        "shortDescription": "11.0\" Dynamic AMOLED 2X, 8400mAh battery, S Pen in box."
    },
    {
        "name": "Samsung Galaxy Tab S9 FE / FE+ (2023)",
        "brand": "Samsung",
        "category": "tablets",
        "basePrice": 1200,
        "colors": "Mint, Silver, Gray, Lavender",
        "variants": [
            {"spec": "128GB / 6GB RAM", "boost": 0},
            {"spec": "256GB / 8GB RAM", "boost": 200},
        ],
        "shortDescription": "10.9\" / 12.4\" display, S Pen inbox, IP68 water resistance."
    },

    # 5. Galaxy Smartwatches
    {
        "name": "Samsung Galaxy Watch 6 Classic (2023)",
        "brand": "Samsung",
        "category": "smartwatches",
        "basePrice": 650,
        "colors": "Black, Silver",
        "variants": [
            {"spec": "43mm (Bluetooth)", "boost": 0},
            {"spec": "43mm (LTE)", "boost": 100},
            {"spec": "47mm (Bluetooth)", "boost": 120},
            {"spec": "47mm (LTE)", "boost": 220},
        ],
        "shortDescription": "Rotating physical bezel, Sapphire Crystal glass, Advanced Sleep Tracking."
    },
    {
        "name": "Samsung Galaxy Watch 6 (2023)",
        "brand": "Samsung",
        "category": "smartwatches",
        "basePrice": 500,
        "colors": "Graphite, Silver, Gold",
        "variants": [
            {"spec": "40mm (Bluetooth)", "boost": 0},
            {"spec": "40mm (LTE)", "boost": 80},
            {"spec": "44mm (Bluetooth)", "boost": 100},
            {"spec": "44mm (LTE)", "boost": 180},
        ],
        "shortDescription": "20% larger display, slimmer bezel, One-Click band system."
    },
    {
        "name": "Samsung Galaxy Watch 5 Pro (2022)",
        "brand": "Samsung",
        "category": "smartwatches",
        "basePrice": 550,
        "colors": "Black Titanium, Gray Titanium",
        "variants": [
            {"spec": "45mm (Bluetooth)", "boost": 0},
            {"spec": "45mm (LTE)", "boost": 120},
        ],
        "shortDescription": "Titanium case, D-Buckle Sport Band, 590mAh battery, Route GPX navigation."
    },

    # 6. Galaxy Laptops (Book Series)
    {
        "name": "Samsung Galaxy Book 4 Ultra (2024)",
        "brand": "Samsung",
        "category": "laptops",
        "basePrice": 5800,
        "colors": "Moonstone Gray",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / RTX 4050", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX 4070", "boost": 1200},
        ],
        "shortDescription": "Intel Core Ultra 7/9 processor, NVIDIA RTX 40-series GPU, 16\" 3K AMOLED Touch."
    },
    {
        "name": "Samsung Galaxy Book 4 Pro 16\" / 14\" (2024)",
        "brand": "Samsung",
        "category": "laptops",
        "basePrice": 4100,
        "colors": "Moonstone Gray, Platinum Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "16GB RAM / 1TB SSD", "boost": 450},
            {"spec": "32GB RAM / 1TB SSD", "boost": 900},
        ],
        "shortDescription": "Intel Core Ultra 7 processor, Dynamic AMOLED 2X Touchscreen, Vision Booster."
    }
]

# 1. Create Excel File
wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=16, bold=True, color="10B981")
bold_font = Font(name="Calibri", size=10, bold=True)
price_font = Font(name="Calibri", size=10, bold=True, color="059669")
border_thin = Side(border_style="thin", color="CBD5E1")
cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

# Sheet 1: Master Catalog
ws1 = wb.active
ws1.title = "Samsung Admin Master Catalog"
ws1.append(["Samsung Electronics - Admin Panel Master Catalog"])
ws1.merge_cells("A1:G1")
ws1["A1"].font = title_font
ws1.append([])

headers1 = [
    "Product Name", "Brand", "Category", "Base Price (AED)", 
    "Color Options", "Specification Variants & Price Boosts", "Short Description"
]
ws1.append(headers1)
for col_num, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx = 4
for p in samsung_products:
    spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
    ws1.append([
        p["name"], p["brand"], p["category"], p["basePrice"],
        p["colors"], spec_list_str, p["shortDescription"]
    ])
    ws1.cell(row=row_idx, column=1).font = bold_font
    ws1.cell(row=row_idx, column=4).font = price_font
    ws1.cell(row=row_idx, column=4).number_format = '#,##0'
    for c in range(1, 8):
        cell = ws1.cell(row=row_idx, column=c)
        cell.border = cell_border
        if c in [2, 3]: cell.alignment = Alignment(horizontal="center")
        elif c == 4: cell.alignment = Alignment(horizontal="right")
        else: cell.alignment = Alignment(horizontal="left", wrap_text=True)
    row_idx += 1

# Sheet 2: Itemized Variants
ws2 = wb.create_sheet(title="Itemized Spec Variants")
ws2.append(["Samsung Specifications & Buyback Valuations Breakdown"])
ws2.merge_cells("A1:H1")
ws2["A1"].font = title_font
ws2.append([])

headers2 = [
    "Product Model Name", "Brand", "Category", "Base Price (AED)",
    "Variant Spec / UOM", "Spec Boost (AED)", "Total Valuation (AED)", "Colors"
]
ws2.append(headers2)
for col_num, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx2 = 4
for p in samsung_products:
    for v in p["variants"]:
        total_val = p["basePrice"] + v["boost"]
        ws2.append([
            p["name"], p["brand"], p["category"], p["basePrice"],
            v["spec"], v["boost"], total_val, p["colors"]
        ])
        ws2.cell(row=row_idx2, column=1).font = bold_font
        ws2.cell(row=row_idx2, column=5).font = bold_font
        ws2.cell(row=row_idx2, column=7).font = price_font
        for c in range(1, 9):
            cell = ws2.cell(row=row_idx2, column=c)
            cell.border = cell_border
            if c in [2, 3]: cell.alignment = Alignment(horizontal="center")
            elif c in [4, 6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            else: cell.alignment = Alignment(horizontal="left")
        row_idx2 += 1

for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

# Save files to workspace and artifact directories
dirs = [r"c:\Users\shant\sellphone.ae", r"C:\Users\shant\.gemini\antigravity-ide\brain\ceae0dbf-d5b8-4d55-b6c1-296d4da1afdd"]

for d in dirs:
    os.makedirs(d, exist_ok=True)
    wb.save(os.path.join(d, "Samsung_Devices_Catalog.xlsx"))
    
    # Write CSV 1: Main Catalog
    with open(os.path.join(d, "Samsung_Devices_Catalog.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers1)
        for p in samsung_products:
            spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
            writer.writerow([p["name"], p["brand"], p["category"], p["basePrice"], p["colors"], spec_list_str, p["shortDescription"]])

    # Write CSV 2: Itemized Variants
    with open(os.path.join(d, "Samsung_Spec_Variants_Breakdown.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers2)
        for p in samsung_products:
            for v in p["variants"]:
                writer.writerow([p["name"], p["brand"], p["category"], p["basePrice"], v["spec"], v["boost"], p["basePrice"] + v["boost"], p["colors"]])

print("Successfully generated Samsung Excel and CSV files!")
