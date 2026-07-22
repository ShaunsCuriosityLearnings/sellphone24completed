import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

macbook_products = [
    {
        "name": "MacBook Air 13\" M3 (2024)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 3200,
        "colors": "Midnight, Starlight, Space Gray, Silver",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "8GB RAM / 512GB SSD", "boost": 300},
            {"spec": "16GB RAM / 512GB SSD", "boost": 550},
            {"spec": "16GB RAM / 1TB SSD", "boost": 850},
            {"spec": "24GB RAM / 1TB SSD", "boost": 1200},
        ],
        "shortDescription": "Apple M3 chip with 8-core CPU, 8-core or 10-core GPU, 13.6\" Liquid Retina display."
    },
    {
        "name": "MacBook Air 15\" M3 (2024)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 3800,
        "colors": "Midnight, Starlight, Space Gray, Silver",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 600},
            {"spec": "16GB RAM / 1TB SSD", "boost": 900},
            {"spec": "24GB RAM / 2TB SSD", "boost": 1600},
        ],
        "shortDescription": "Apple M3 chip, 15.3\" Liquid Retina display, 10-core GPU, 6-speaker sound system."
    },
    {
        "name": "MacBook Air 13\" M2 (2022)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 2400,
        "colors": "Midnight, Starlight, Space Gray, Silver",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "8GB RAM / 512GB SSD", "boost": 250},
            {"spec": "16GB RAM / 512GB SSD", "boost": 450},
            {"spec": "16GB RAM / 1TB SSD", "boost": 750},
            {"spec": "24GB RAM / 1TB SSD", "boost": 1050},
        ],
        "shortDescription": "Apple M2 chip with 8-core CPU, 8/10-core GPU, 13.6\" Liquid Retina display, MagSafe 3."
    },
    {
        "name": "MacBook Air 15\" M2 (2023)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 2900,
        "colors": "Midnight, Starlight, Space Gray, Silver",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 500},
            {"spec": "16GB RAM / 1TB SSD", "boost": 800},
            {"spec": "24GB RAM / 1TB SSD", "boost": 1100},
        ],
        "shortDescription": "Apple M2 chip, 15.3\" Liquid Retina display, 10-core GPU, fanless design."
    },
    {
        "name": "MacBook Air 13\" M1 (2020)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 1600,
        "colors": "Space Gray, Silver, Gold",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "8GB RAM / 512GB SSD", "boost": 200},
            {"spec": "16GB RAM / 512GB SSD", "boost": 350},
            {"spec": "16GB RAM / 1TB SSD", "boost": 600},
        ],
        "shortDescription": "Apple M1 chip with 8-core CPU, 7/8-core GPU, 13.3\" Retina display, fanless design."
    },
    {
        "name": "MacBook Pro 14\" M3 (2023)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 4200,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "8GB RAM / 512GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 400},
            {"spec": "16GB RAM / 1TB SSD", "boost": 750},
            {"spec": "24GB RAM / 1TB SSD", "boost": 1100},
        ],
        "shortDescription": "Apple M3 chip with 8-core CPU, 10-core GPU, 14.2\" Liquid Retina XDR 120Hz display."
    },
    {
        "name": "MacBook Pro 14\" M3 Pro / Max (2023)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 5200,
        "colors": "Space Black, Silver",
        "variants": [
            {"spec": "18GB RAM / 512GB SSD", "boost": 0},
            {"spec": "18GB RAM / 1TB SSD", "boost": 500},
            {"spec": "36GB RAM / 1TB SSD", "boost": 1100},
            {"spec": "48GB RAM / 1TB SSD", "boost": 1600},
            {"spec": "64GB RAM / 2TB SSD", "boost": 2400},
            {"spec": "128GB RAM / 4TB SSD", "boost": 4200},
        ],
        "shortDescription": "Apple M3 Pro/Max chip, 14.2\" Liquid Retina XDR display, HDMI, SDXC, MagSafe 3."
    },
    {
        "name": "MacBook Pro 16\" M3 Pro / Max (2023)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 6400,
        "colors": "Space Black, Silver",
        "variants": [
            {"spec": "18GB RAM / 512GB SSD", "boost": 0},
            {"spec": "36GB RAM / 512GB SSD", "boost": 700},
            {"spec": "36GB RAM / 1TB SSD", "boost": 1100},
            {"spec": "48GB RAM / 1TB SSD", "boost": 1700},
            {"spec": "64GB RAM / 2TB SSD", "boost": 2600},
            {"spec": "128GB RAM / 4TB SSD", "boost": 4500},
        ],
        "shortDescription": "Apple M3 Pro/Max chip, 16.2\" Liquid Retina XDR display, up to 40-core GPU."
    },
    {
        "name": "MacBook Pro 14\" M2 Pro / Max (2023)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 4100,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "16GB RAM / 1TB SSD", "boost": 450},
            {"spec": "32GB RAM / 1TB SSD", "boost": 950},
            {"spec": "64GB RAM / 2TB SSD", "boost": 2000},
            {"spec": "96GB RAM / 4TB SSD", "boost": 3600},
        ],
        "shortDescription": "Apple M2 Pro/Max chip, 14.2\" Liquid Retina XDR display, 10/12-core CPU, up to 38-core GPU."
    },
    {
        "name": "MacBook Pro 16\" M2 Pro / Max (2023)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 5100,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "32GB RAM / 512GB SSD", "boost": 600},
            {"spec": "32GB RAM / 1TB SSD", "boost": 1000},
            {"spec": "64GB RAM / 2TB SSD", "boost": 2200},
            {"spec": "96GB RAM / 4TB SSD", "boost": 3800},
        ],
        "shortDescription": "Apple M2 Pro/Max chip, 16.2\" Liquid Retina XDR display, HDMI 2.1, Wi-Fi 6E."
    },
    {
        "name": "MacBook Pro 14\" M1 Pro / Max (2021)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 3200,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "16GB RAM / 1TB SSD", "boost": 400},
            {"spec": "32GB RAM / 1TB SSD", "boost": 800},
            {"spec": "64GB RAM / 2TB SSD", "boost": 1700},
        ],
        "shortDescription": "Apple M1 Pro/Max chip, 14.2\" ProMotion 120Hz Liquid Retina XDR display, MagSafe 3."
    },
    {
        "name": "MacBook Pro 16\" M1 Pro / Max (2021)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 4000,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "32GB RAM / 512GB SSD", "boost": 500},
            {"spec": "32GB RAM / 1TB SSD", "boost": 850},
            {"spec": "64GB RAM / 2TB SSD", "boost": 1900},
        ],
        "shortDescription": "Apple M1 Pro/Max chip, 16.2\" ProMotion Liquid Retina XDR display, 10-core CPU."
    },
    {
        "name": "MacBook Pro 13\" M2 (2022)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 2200,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "8GB RAM / 512GB SSD", "boost": 250},
            {"spec": "16GB RAM / 512GB SSD", "boost": 450},
            {"spec": "24GB RAM / 1TB SSD", "boost": 900},
        ],
        "shortDescription": "Apple M2 chip with Touch Bar, 13.3\" Retina display, active cooling fan."
    },
    {
        "name": "MacBook Pro 13\" M1 (2020)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 1700,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "8GB RAM / 512GB SSD", "boost": 200},
            {"spec": "16GB RAM / 512GB SSD", "boost": 350},
            {"spec": "16GB RAM / 1TB SSD", "boost": 600},
        ],
        "shortDescription": "Apple M1 chip with Touch Bar, 13.3\" Retina display, active cooling fan."
    },
    {
        "name": "MacBook Pro 16\" Intel i7 / i9 (2019)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 1400,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "16GB RAM / 1TB SSD", "boost": 200},
            {"spec": "32GB RAM / 1TB SSD", "boost": 450},
            {"spec": "64GB RAM / 2TB SSD", "boost": 800},
        ],
        "shortDescription": "Intel Core i7/i9 processor, AMD Radeon Pro GPU, 16.0\" Retina display."
    },
    {
        "name": "MacBook Air 13\" Intel i3 / i5 / i7 (2020)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 1100,
        "colors": "Space Gray, Silver, Gold",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "8GB RAM / 512GB SSD", "boost": 150},
            {"spec": "16GB RAM / 512GB SSD", "boost": 300},
        ],
        "shortDescription": "10th Gen Intel Core i3/i5/i7 processor, Magic Keyboard, 13.3\" Retina display."
    },
    {
        "name": "MacBook Pro 13\" Intel i5 / i7 (2020 - 4 TB3 Ports)",
        "brand": "Apple",
        "category": "laptops",
        "basePrice": 1250,
        "colors": "Space Gray, Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "16GB RAM / 1TB SSD", "boost": 250},
            {"spec": "32GB RAM / 2TB SSD", "boost": 600},
        ],
        "shortDescription": "10th Gen Intel Core i5/i7 processor, 4 Thunderbolt 3 ports, Touch Bar, Magic Keyboard."
    }
]

wb = openpyxl.Workbook()

# Setup styles
header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=16, bold=True, color="10B981")
bold_font = Font(name="Calibri", size=10, bold=True)
regular_font = Font(name="Calibri", size=10)
price_font = Font(name="Calibri", size=10, bold=True, color="059669")
border_thin = Side(border_style="thin", color="CBD5E1")
cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

# Sheet 1: Master Admin Products
ws1 = wb.active
ws1.title = "Admin Product Master Catalog"

ws1.append(["Apple MacBook Laptops - Admin Panel Master Catalog"])
ws1.merge_cells("A1:G1")
ws1["A1"].font = title_font

ws1.append([])

headers1 = [
    "Product Name", "Brand", "Category", "Base Price (AED)", 
    "Color Options", "Specification Variants & Price Boosts (JSON/List)", "Short Description"
]
ws1.append(headers1)

for col_num, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx = 4
for p in macbook_products:
    spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
    ws1.append([
        p["name"],
        p["brand"],
        p["category"],
        p["basePrice"],
        p["colors"],
        spec_list_str,
        p["shortDescription"]
    ])
    
    ws1.cell(row=row_idx, column=1).font = bold_font
    ws1.cell(row=row_idx, column=4).font = price_font
    ws1.cell(row=row_idx, column=4).number_format = '#,##0'
    
    for c in range(1, 8):
        cell = ws1.cell(row=row_idx, column=c)
        cell.border = cell_border
        if c in [2, 3]:
            cell.alignment = Alignment(horizontal="center")
        elif c == 4:
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left", wrap_text=True)
    row_idx += 1

# Sheet 2: Itemized Variant Breakdown
ws2 = wb.create_sheet(title="Itemized Spec Variants")

ws2.append(["Apple MacBook Specifications & Buyback Valuations Breakdown"])
ws2.merge_cells("A1:H1")
ws2["A1"].font = title_font
ws2.append([])

headers2 = [
    "Product Model Name", "Brand", "Category", "Base Price (AED)",
    "Variant Spec / UOM (RAM & Storage)", "Spec Boost (AED)", "Total Valuation (AED)", "Colors"
]
ws2.append(headers2)

for col_num, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx2 = 4
for p in macbook_products:
    for v in p["variants"]:
        total_val = p["basePrice"] + v["boost"]
        ws2.append([
            p["name"],
            p["brand"],
            p["category"],
            p["basePrice"],
            v["spec"],
            v["boost"],
            total_val,
            p["colors"]
        ])
        
        ws2.cell(row=row_idx2, column=1).font = bold_font
        ws2.cell(row=row_idx2, column=5).font = bold_font
        ws2.cell(row=row_idx2, column=7).font = price_font
        
        for c in range(1, 9):
            cell = ws2.cell(row=row_idx2, column=c)
            cell.border = cell_border
            if c in [2, 3]:
                cell.alignment = Alignment(horizontal="center")
            elif c in [4, 6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left")
        row_idx2 += 1

# Auto-adjust column widths for both sheets
for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

# Save output files
paths = [
    r"c:\Users\shant\sellphone.ae\Apple_MacBook_Laptops_Catalog.xlsx",
    r"C:\Users\shant\.gemini\antigravity-ide\brain\ceae0dbf-d5b8-4d55-b6c1-296d4da1afdd\Apple_MacBook_Laptops_Catalog.xlsx"
]

for path in paths:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"Saved Excel file to {path}")
