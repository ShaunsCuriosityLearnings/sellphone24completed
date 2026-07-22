import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

dell_laptops = [
    # 1. Dell XPS Premium Series
    {
        "name": "Dell XPS 16 9640 (2024)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 4800,
        "colors": "Platinum, Graphite",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / RTX 4050", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX 4060", "boost": 800},
            {"spec": "64GB RAM / 2TB SSD / RTX 4070", "boost": 1800}
        ],
        "shortDescription": "Intel Core Ultra 7/9, 16.3\" 4K+ OLED Touchscreen, CNC Machined Aluminum, RTX 40-series GPU."
    },
    {
        "name": "Dell XPS 14 9440 (2024)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 3900,
        "colors": "Platinum, Graphite",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / Intel Arc", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX 4050", "boost": 700},
            {"spec": "64GB RAM / 2TB SSD / RTX 4050", "boost": 1400}
        ],
        "shortDescription": "Intel Core Ultra 7, 14.5\" 3.2K OLED Touch, Seamless Glass Touchpad, Gorilla Glass Victus."
    },
    {
        "name": "Dell XPS 15 9530 (2023)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 3400,
        "colors": "Platinum Silver, Carbon Fiber Black",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / RTX 4050", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX 4060", "boost": 600},
            {"spec": "64GB RAM / 2TB SSD / RTX 4070", "boost": 1300}
        ],
        "shortDescription": "13th Gen Intel Core i7/i9, 15.6\" 3.5K OLED Touch, Quad Speaker System, Carbon Fiber Palm Rest."
    },
    {
        "name": "Dell XPS 17 9730 (2023)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 4100,
        "colors": "Platinum Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / RTX 4060", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX 4070", "boost": 750},
            {"spec": "64GB RAM / 2TB SSD / RTX 4080", "boost": 1700}
        ],
        "shortDescription": "13th Gen Intel Core i7/i9, 17.0\" 4K UHD+ Touch, Vapor Chamber Cooling, 4x Thunderbolt 4."
    },
    {
        "name": "Dell XPS 13 Plus 9320 (2022-2023)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 2800,
        "colors": "Platinum, Graphite",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD", "boost": 500},
            {"spec": "32GB RAM / 2TB SSD", "boost": 900}
        ],
        "shortDescription": "12th/13th Gen Intel Core i7, Capacitive Touch Function Row, Zero-Lattice Keyboard, 3.5K OLED."
    },
    {
        "name": "Dell XPS 15 9500 / 9510 / 9520 (2020-2022)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 2200,
        "colors": "Platinum Silver, Frost White",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / GTX 1650 Ti", "boost": 0},
            {"spec": "16GB RAM / 1TB SSD / RTX 3050 Ti", "boost": 400},
            {"spec": "32GB RAM / 1TB SSD / RTX 3050 Ti", "boost": 750}
        ],
        "shortDescription": "10th-12th Gen Intel Core i7/i9, 15.6\" 4K UHD+ Touch display, Waves MaxxAudio Pro."
    },
    {
        "name": "Dell XPS 13 9300 / 9310 (2020-2021)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 1600,
        "colors": "Platinum Silver, Arctic White",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 300},
            {"spec": "16GB RAM / 1TB SSD", "boost": 500}
        ],
        "shortDescription": "10th/11th Gen Intel Core i5/i7, 13.4\" 4K UHD+ InfinityEdge Display, Thunderbolt 4."
    },
    {
        "name": "Dell XPS 15 9560 / 9570 / 7590 (2017-2019)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 1400,
        "colors": "Silver, Black Carbon",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / GTX 1050 Ti", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / GTX 1650", "boost": 350}
        ],
        "shortDescription": "7th-9th Gen Intel Core i7/i9, 15.6\" 4K OLED/4K Touch Display, NVIDIA GTX Graphics."
    },
    {
        "name": "Dell XPS 13 9343 / 9350 / 9360 / 9370 (2015-2018)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 950,
        "colors": "Silver, Rose Gold",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 200}
        ],
        "shortDescription": "5th-8th Gen Intel Core i5/i7, 13.3\" QHD+ InfinityEdge Touch screen, Machined Aluminum."
    },
    {
        "name": "Dell XPS 15 L501X / L502X / 9530 (2010-2014)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 550,
        "colors": "Anodized Aluminum Silver",
        "variants": [
            {"spec": "8GB RAM / 500GB HDD", "boost": 0},
            {"spec": "16GB RAM / 256GB SSD", "boost": 150}
        ],
        "shortDescription": "Intel 1st-4th Gen Core i5/i7, JBL Speakers with Subwoofer, NVIDIA GeForce GT Graphics."
    },

    # 2. Dell Alienware Gaming Series
    {
        "name": "Dell Alienware m16 / m18 R2 (2024)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 5200,
        "colors": "Dark Metallic Moon",
        "variants": [
            {"spec": "16GB RAM / 1TB SSD / RTX 4070", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX 4080", "boost": 1100},
            {"spec": "64GB RAM / 2TB SSD / RTX 4090", "boost": 2400}
        ],
        "shortDescription": "14th Gen Intel Core i9-14900HX, 240Hz QHD+ Display, Element 31 Thermal Interface, CherryMX Mechanical Keyboard."
    },
    {
        "name": "Dell Alienware x16 / x14 R2 (2023)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 4200,
        "colors": "Lunar Light",
        "variants": [
            {"spec": "16GB RAM / 1TB SSD / RTX 4060", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX 4070", "boost": 800},
            {"spec": "32GB RAM / 2TB SSD / RTX 4080", "boost": 1600}
        ],
        "shortDescription": "Ultra-thin gaming laptop, 13th Gen Intel Core i7/i9, Legend 3.0 design, AlienFX RGB Lighting."
    },
    {
        "name": "Dell Alienware m15 R6 / R7 (2021-2022)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 2800,
        "colors": "Dark Side of the Moon, Lunar Light",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / RTX 3060", "boost": 0},
            {"spec": "16GB RAM / 1TB SSD / RTX 3070 Ti", "boost": 500},
            {"spec": "32GB RAM / 1TB SSD / RTX 3080", "boost": 1000}
        ],
        "shortDescription": "11th/12th Gen Intel Core i7/i9, 15.6\" 240Hz/360Hz Gaming Display, Cryo-Tech Cooling."
    },
    {
        "name": "Dell Alienware 17 R4 / R5 (2016-2018)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 1600,
        "colors": "Epic Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / GTX 1070", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / GTX 1080", "boost": 400}
        ],
        "shortDescription": "7th/8th Gen Intel Core i7/i9, Tobii Eye Tracking, 17.3\" QHD 120Hz, NVIDIA GTX 10-series."
    },
    {
        "name": "Dell Alienware M17x R3 / R4 / 18 (2011-2015)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 850,
        "colors": "Stealth Black, Nebula Red",
        "variants": [
            {"spec": "12GB RAM / 750GB HDD / GTX 675M", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD / GTX 880M SLI", "boost": 300}
        ],
        "shortDescription": "Intel 2nd-4th Gen Core i7, Dual SLI Graphics, 17.3\" / 18.4\" Full HD display, Anodized Aluminum chassis."
    },

    # 3. Dell Latitude Business Series
    {
        "name": "Dell Latitude 7440 / 7430 (2022-2023)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 2400,
        "colors": "Titan Gray, Aluminum Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD", "boost": 450}
        ],
        "shortDescription": "12th/13th Gen Intel Core i5/i7 vPro, 14.0\" FHD+ 16:10 display, Magnesium/Aluminum Chassis, Wi-Fi 6E."
    },
    {
        "name": "Dell Latitude 5440 / 5430 (2022-2023)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 1800,
        "colors": "Gray",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 250},
            {"spec": "32GB RAM / 1TB SSD", "boost": 500}
        ],
        "shortDescription": "12th/13th Gen Intel Core i5/i7, 14.0\" Full HD, Bio-based materials, ExpressCharge."
    },
    {
        "name": "Dell Latitude 7400 / 7490 / 7480 (2017-2019)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 1100,
        "colors": "Black Carbon Fiber",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 200}
        ],
        "shortDescription": "7th/8th Gen Intel Core i5/i7, 14.0\" FHD Anti-Glare display, SmartCard Reader, Military-Grade Durability."
    },
    {
        "name": "Dell Latitude E7470 / E7450 / E7440 (2014-2016)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 650,
        "colors": "Black",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 150}
        ],
        "shortDescription": "4th-6th Gen Intel Core i5/i7, Tri-metal alloy chassis, 14.0\" Full HD, Docking connector."
    },
    {
        "name": "Dell Latitude E6430 / E6420 / E6410 (2010-2013)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 400,
        "colors": "Brushed Aluminum Gray",
        "variants": [
            {"spec": "4GB RAM / 320GB HDD", "boost": 0},
            {"spec": "8GB RAM / 256GB SSD", "boost": 120}
        ],
        "shortDescription": "Intel 1st-3rd Gen Core i5/i7, Tri-Metal Casing, DVD-RW drive, Spill-Resistant Keyboard."
    },

    # 4. Dell Inspiron & G-Series Gaming
    {
        "name": "Dell G15 5530 / G16 7630 Gaming (2023-2024)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 2900,
        "colors": "Dark Shadow Gray, Quantum White",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / RTX 4050", "boost": 0},
            {"spec": "16GB RAM / 1TB SSD / RTX 4060", "boost": 450},
            {"spec": "32GB RAM / 1TB SSD / RTX 4070", "boost": 950}
        ],
        "shortDescription": "13th Gen Intel Core i7/i9, 15.6\"/16.0\" 165Hz/240Hz Display, Alienware-inspired Cooling."
    },
    {
        "name": "Dell Inspiron 16 Plus 7630 / 7620 (2022-2023)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 2200,
        "colors": "Dark Green, Ice Blue",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / RTX 3050", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX 4060", "boost": 600}
        ],
        "shortDescription": "12th/13th Gen Intel Core i7 H-Series, 16.0\" 2.5K/3K 120Hz display, Quad Speakers."
    },
    {
        "name": "Dell Inspiron 15 3511 / 5510 / 5502 (2020-2022)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 1200,
        "colors": "Carbon Black, Platinum Silver",
        "variants": [
            {"spec": "8GB RAM / 256GB SSD", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD", "boost": 200}
        ],
        "shortDescription": "11th/12th Gen Intel Core i5/i7, 15.6\" FHD Anti-Glare display, Lift Hinge design."
    },
    {
        "name": "Dell Inspiron 15 7000 Gaming 7567 / 7577 (2016-2018)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 950,
        "colors": "Matte Black, Beijing Red",
        "variants": [
            {"spec": "8GB RAM / 128GB SSD + 1TB HDD / GTX 1050", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD / GTX 1060 Max-Q", "boost": 250}
        ],
        "shortDescription": "7th/8th Gen Intel Core i5/i7, Dual Fan Thermal Cooling, Red Backlit Keyboard, NVIDIA GTX graphics."
    },
    {
        "name": "Dell Inspiron 15R 5520 / N5110 / N5010 (2010-2013)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 350,
        "colors": "Switch Changeable Lids (Black, Red, Blue)",
        "variants": [
            {"spec": "4GB RAM / 500GB HDD", "boost": 0},
            {"spec": "8GB RAM / 246GB SSD", "boost": 100}
        ],
        "shortDescription": "Intel 1st-3rd Gen Core i3/i5/i7, 15.6\" HD TruLife display, SWITCH by Design Studio interchangeable covers."
    },

    # 5. Dell Precision Workstations
    {
        "name": "Dell Precision 5680 / 5570 Mobile Workstation (2022-2023)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 4200,
        "colors": "Titan Gray",
        "variants": [
            {"spec": "32GB RAM / 512GB SSD / RTX A2000", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / RTX A3500", "boost": 800},
            {"spec": "64GB RAM / 2TB SSD / RTX 5000 Ada", "boost": 2100}
        ],
        "shortDescription": "ISV Certified Professional Workstation, 13th Gen Intel Core i7/i9, 16.0\" 4K OLED Touch, NVIDIA RTX Ada Graphics."
    },
    {
        "name": "Dell Precision 7550 / 7530 Workstation (2018-2020)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 2100,
        "colors": "Aluminum Silver",
        "variants": [
            {"spec": "16GB RAM / 512GB SSD / Quadro T2000", "boost": 0},
            {"spec": "32GB RAM / 1TB SSD / Quadro RTX 4000", "boost": 650},
            {"spec": "64GB RAM / 2TB SSD / Quadro RTX 5000", "boost": 1300}
        ],
        "shortDescription": "8th-10th Gen Intel Core i7/i9 / Xeon, Up to 128GB RAM support, 15.6\" 4K Adobe RGB Display."
    },
    {
        "name": "Dell Precision M4800 / M4600 Workstation (2011-2014)",
        "brand": "Dell",
        "category": "laptops",
        "basePrice": 750,
        "colors": "Coventry Brown Metallic",
        "variants": [
            {"spec": "8GB RAM / 500GB HDD / Quadro K1100M", "boost": 0},
            {"spec": "16GB RAM / 512GB SSD / Quadro K2100M", "boost": 200}
        ],
        "shortDescription": "Intel 2nd-4th Gen Core i7 / Extreme, 15.6\" QHD+ 3200x1800 UltraSharp display, Magnesium Alloy chassis."
    }
]

# Generate Excel Workbook
wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=16, bold=True, color="10B981")
bold_font = Font(name="Calibri", size=10, bold=True)
price_font = Font(name="Calibri", size=10, bold=True, color="059669")
border_thin = Side(border_style="thin", color="CBD5E1")
cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

# Sheet 1: Master Catalog
ws1 = wb.active
ws1.title = "Dell Laptops Admin Master Catalog"
ws1.append(["Dell Laptops Catalog (2010 - 2024) - Admin Panel Master Catalog"])
ws1.merge_cells("A1:G1")
ws1["A1"].font = title_font
ws1.append([])

headers1 = [
    "Product Name", "Brand", "Category", "Base Price (AED)", 
    "Color Options", "Specification Variants & Price Boosts", "Short Description"
]
ws1.append(headers1)
for col_num, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx = 4
for p in dell_laptops:
    spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
    ws1.append([
        p["name"], p["brand"], p["category"], p["basePrice"],
        p["colors"], spec_list_str, p["shortDescription"]
    ])
    ws1.cell(row=row_idx, column=1).font = bold_font
    ws1.cell(row=row_idx, column=4).font = price_font
    ws1.cell(row=row_idx, column=4).number_format = '#,##0'
    for c in range(1, 8):
        cell = ws1.cell(row=row_idx, column=c)
        cell.border = cell_border
        if c in [2, 3]: cell.alignment = Alignment(horizontal="center")
        elif c == 4: cell.alignment = Alignment(horizontal="right")
        else: cell.alignment = Alignment(horizontal="left", wrap_text=True)
    row_idx += 1

# Sheet 2: Itemized Variants
ws2 = wb.create_sheet(title="Itemized Spec Variants")
ws2.append(["Dell Laptop Specifications & Buyback Valuations Breakdown"])
ws2.merge_cells("A1:H1")
ws2["A1"].font = title_font
ws2.append([])

headers2 = [
    "Product Model Name", "Brand", "Category", "Base Price (AED)",
    "Variant Spec / Processor & RAM & SSD", "Spec Boost (AED)", "Total Valuation (AED)", "Colors"
]
ws2.append(headers2)
for col_num, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

row_idx2 = 4
for p in dell_laptops:
    for v in p["variants"]:
        total_val = p["basePrice"] + v["boost"]
        ws2.append([
            p["name"], p["brand"], p["category"], p["basePrice"],
            v["spec"], v["boost"], total_val, p["colors"]
        ])
        ws2.cell(row=row_idx2, column=1).font = bold_font
        ws2.cell(row=row_idx2, column=5).font = bold_font
        ws2.cell(row=row_idx2, column=7).font = price_font
        for c in range(1, 9):
            cell = ws2.cell(row=row_idx2, column=c)
            cell.border = cell_border
            if c in [2, 3]: cell.alignment = Alignment(horizontal="center")
            elif c in [4, 6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            else: cell.alignment = Alignment(horizontal="left")
        row_idx2 += 1

for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

# Save files to workspace and artifact directories
dirs = [r"c:\Users\shant\sellphone.ae", r"C:\Users\shant\.gemini\antigravity-ide\brain\ceae0dbf-d5b8-4d55-b6c1-296d4da1afdd"]

for d in dirs:
    os.makedirs(d, exist_ok=True)
    wb.save(os.path.join(d, "Dell_Laptops_Catalog.xlsx"))
    
    # Write CSV 1: Main Catalog
    with open(os.path.join(d, "Dell_Laptops_Catalog.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers1)
        for p in dell_laptops:
            spec_list_str = "; ".join([f"{v['spec']} (+AED {v['boost']})" for v in p["variants"]])
            writer.writerow([p["name"], p["brand"], p["category"], p["basePrice"], p["colors"], spec_list_str, p["shortDescription"]])

    # Write CSV 2: Itemized Variants
    with open(os.path.join(d, "Dell_Laptops_Variants_Breakdown.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers2)
        for p in dell_laptops:
            for v in p["variants"]:
                writer.writerow([p["name"], p["brand"], p["category"], p["basePrice"], v["spec"], v["boost"], p["basePrice"] + v["boost"], p["colors"]])

print("Successfully generated Dell Laptops Excel and CSV files!")
